import os
import hashlib   # this library is used to encrypt the password in hashkey
from tkinter  import *     # for gui (Graphical User Interface)
from tkinter import messagebox
from openpyxl import Workbook , load_workbook      # for manipulating the excel file

file_path = "C:/Users/sigde/OneDrive/Desktop/Python/Major_Projects/Users.xlsx"
def encrypt_password(password):
    encrypted_password =  hashlib.sha256(password.encode()).hexdigest()
    return encrypted_password

# def create_file():
#     if not os.path.exists(file_path):
#         wb = Workbook()
#         sheet = wb.active
#         sheet.title = "Users"
#         sheet.append("Username","Password")
#         wb.save(file_path)

def register_user():
    
    username = entry_username.get()
    password = entry_password.get()
    
    if username == "" or password == "":
        print("Empty username or password")
        # messagebox.showerror("Error","All fields are required")
        
    encrypted_password = encrypt_password(password)
    wb = load_workbook(file_path)
    sheet = wb.active
    for row in sheet.iter_rows(min_row = 2, values_only= True):
        if row[0] == username:
            print("User name already exist")
            #...... gui ko logic
    sheet.append(username,encrypted_password)
    wb.save(file_path)
    
    # gui ko logic